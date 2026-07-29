#!/usr/bin/env python3
"""
wk_floor_review.py : merge the founder's floor-review workbook
(WK_Floor_Review.xlsx) into the standards store, through the tested
loader, without teaching the loader a one-time review vocabulary
(ROUND6_CLOSEOUT session R).

The workbook carries three review tabs ("Floors (300)", "Method sweep A
(41)", "Method sweep B (130)"), each with a DECISION column holding a
controlled vocabulary: keep / floor_1 / floor_2 / process / method /
preference / unsure, with blank meaning NOT YET REVIEWED (distinct from
keep). "Read me first" and "Open decisions" tabs are founder prose and
are ignored. Columns are found by header name, not position.

The three rules (verbatim from the brief):
  1. unsure never imports. It is a queue, not a decision; those rows are
     listed with their current tier and text and left untouched.
  2. Blank is not keep. Blank counts are reported per tab, and a partial
     import requires --allow-partial so it can never happen by accident.
  3. seed_reviewed flips ONLY via the separate --flip-reviewed action,
     and only at zero blanks and zero unsure across all three tabs.

Round seven, session V: a DECIDED row whose store tier differs from the
tier the workbook displayed was decided against stale information. The
script refuses those rows unless --accept-drift, listing each one, so a
decision made against a tier that changed after the sheet was built is
never applied silently. (The undecided half of this class was fixed in
session R: undecided rows always carry the store's current tier.)

Round seven, the review-stamp decision (founder, 2026-07-28): every
committing run emits a dated review record (frontmatter status: frozen)
listing who reviewed, when, what moved, what was confirmed, what stayed
queued. The record lands under docs/ and the script prints the manifest
line for frozen-records.test.ts; the commit that adds the record adds
the hash, which the guard enforces. --reviewer is required on any run
that writes.

Every tier change writes through packages/schema/src/load-provisions.ts
(--supersede), so provision_versions records each move the same as any
other provision edit. This script never writes the database itself.

Dry run by default: it prints the plan and writes the merged seed JSON,
but does not invoke the loader without --commit.

DATABASE_URL is taken from the environment, or read from .neon-connection
at the repo root if present. It is never printed.

Session AN (2026-07-29): the founder's review happened in the review APP,
not the workbook, so the script also reads flat TSV exports (header
provision_id / seen_tier / DECISION / NOTE), one or more at once:

  python3 wk_floor_review.py floor_review_export.tsv tier_corrections.tsv

The AN rules, on top of R and V:
  - seen_tier is REQUIRED in TSV inputs, header and per row; it feeds the
    session V drift check, and a row without it is refused, never assumed.
  - Two inputs disagreeing about one provision is a CONFLICT: refuse,
    list every conflicting provision with both values and sources, stop.
    The founder resolves it; the script never picks.
  - "split" reaching the script is an export bug (the app expands split
    clusters into per-provision rows); refused loudly, never handled.
  - Completeness is redefined against the provision set, not tabs: the
    flip gate requires zero unsure AND no store FLOOR provision left
    without a decision (a floor that went unreviewed blocks the flip; a
    non-floor row the review never included does not). The total
    no-decision count is always reported, since a flat export cannot
    distinguish "not reviewed" from "not included".

Usage:
  python3 wk_floor_review.py WK_Floor_Review.xlsx                # plan only
  python3 wk_floor_review.py export.tsv corrections.tsv          # plan only
  python3 wk_floor_review.py export.tsv corrections.tsv --commit --reviewer "..."
  python3 wk_floor_review.py export.tsv corrections.tsv --flip-reviewed --reviewer "..."
"""

import argparse
import csv
import json
import os
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
DEFAULT_BASE = os.path.join(ROOT, "tooling", "seed", "provisions_seed.json")
LOADER = os.path.join(ROOT, "packages", "schema", "src", "load-provisions.ts")

TIERS = {"floor_1", "floor_2", "process", "method", "preference"}
DECISIONS = TIERS | {"keep", "unsure"}
REVIEW_TABS = ("Floors (300)", "Method sweep A (41)", "Method sweep B (130)")
IGNORED_TABS = ("Read me first", "Open decisions")


def find_columns(ws):
    """Locate provision_id, DECISION, and (optionally) the displayed tier
    by header name (row 1), any case. The tier column powers the session V
    drift check; a tab without one is reported, not failed."""
    cols = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip().lower()
        if h in ("provision_id", "provision id"):
            cols["id"] = c
        elif h == "decision":
            cols["decision"] = c
        elif h in ("tier", "current tier", "current_tier"):
            cols["tier"] = c
    missing = [k for k in ("id", "decision") if k not in cols]
    if missing:
        sys.exit(f"FAIL: tab '{ws.title}' is missing header(s): "
                 f"{', '.join('provision_id' if m == 'id' else 'DECISION' for m in missing)}")
    return cols


def read_tsvs(paths):
    """Session AN: read one or more flat TSV exports (the review app's
    format). Returns the same decision map shape as read_workbook, with
    per-source stats, plus conflicts (same pid, different decisions across
    inputs; the founder resolves, never the script) and per-pid notes.
    seen_tier is required, header and row; split is an export bug."""
    rows_by_pid = {}
    per_source, errors = {}, []
    for path in paths:
        src = os.path.basename(path)
        stats = {"rows": 0, "keep": 0, "change": 0, "unsure": 0, "blank": 0}
        with open(path, newline="") as f:
            reader = csv.reader(f, delimiter="\t")
            try:
                header = [h.strip().lower() for h in next(reader)]
            except StopIteration:
                errors.append(f"{src}: empty file")
                continue
            idx = {name: header.index(name) for name in header}
            missing_cols = [c for c in ("provision_id", "seen_tier", "decision") if c not in idx]
            if missing_cols:
                errors.append(
                    f"{src}: missing column(s) {', '.join(missing_cols)}. seen_tier is required "
                    f"(session V's drift check runs on it); a row without it is refused, never assumed.")
                per_source[src] = stats
                continue
            for n, row in enumerate(reader, start=2):
                if not row or not "".join(row).strip():
                    continue
                def cell(name):
                    i = idx[name]
                    return row[i].strip() if i < len(row) else ""
                pid = cell("provision_id")
                if not pid:
                    continue
                stats["rows"] += 1
                raw = cell("decision").lower()
                seen = cell("seen_tier").lower()
                note = cell("note") if "note" in idx else ""
                if raw == "split":
                    errors.append(f"{src} row {n}: {pid}: DECISION 'split' reached the script; the "
                                  f"app expands split clusters into per-provision rows, so this is "
                                  f"an EXPORT BUG to report, not a value to handle.")
                    continue
                if raw == "":
                    errors.append(f"{src} row {n}: {pid}: blank DECISION; the app omits undecided "
                                  f"rows entirely, so a blank row is malformed, not unreviewed.")
                    continue
                if raw not in DECISIONS:
                    errors.append(f"{src} row {n}: {pid}: DECISION {raw!r} is not in the vocabulary "
                                  f"{sorted(DECISIONS)}")
                    continue
                if not seen:
                    errors.append(f"{src} row {n}: {pid}: seen_tier is blank; refused rather than "
                                  f"assumed (the drift check needs the tier the reviewer SAW).")
                    continue
                rows_by_pid.setdefault(pid, []).append((raw, src, seen, note))
                stats["unsure" if raw == "unsure" else "keep" if raw == "keep" else "change"] += 1
        per_source[src] = stats
    decisions, conflicts, notes = {}, [], {}
    for pid, entries in rows_by_pid.items():
        values = sorted({e[0] for e in entries})
        if len(values) > 1:
            conflicts.append({"provision_id": pid,
                              "values": [f"{e[0]} ({e[1]}, saw {e[2]})" for e in entries]})
            continue
        seens = [e[2] for e in entries]
        sources = ",".join(sorted({e[1] for e in entries}))
        decisions[pid] = (entries[0][0], sources, seens)
        note = next((e[3] for e in entries if e[3]), "")
        if note:
            notes[pid] = note
    return decisions, per_source, errors, conflicts, notes


def read_workbook(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required for workbook input: pip install openpyxl --break-system-packages")
    wb = openpyxl.load_workbook(path, data_only=True)
    present = [t for t in REVIEW_TABS if t in wb.sheetnames]
    if len(present) != len(REVIEW_TABS):
        absent = [t for t in REVIEW_TABS if t not in wb.sheetnames]
        sys.exit(f"FAIL: workbook is missing review tab(s): {', '.join(absent)}. "
                 f"Found: {', '.join(wb.sheetnames)}")
    decisions, errors = {}, []
    per_tab = {}
    no_tier_tabs = []
    for tab in REVIEW_TABS:
        ws = wb[tab]
        cols = find_columns(ws)
        if "tier" not in cols:
            no_tier_tabs.append(tab)
        stats = {"rows": 0, "keep": 0, "change": 0, "unsure": 0, "blank": 0}
        for r in range(2, ws.max_row + 1):
            pid = ws.cell(r, cols["id"]).value
            if not pid:
                continue
            pid = str(pid).strip()
            stats["rows"] += 1
            raw = str(ws.cell(r, cols["decision"]).value or "").strip().lower()
            shown = (str(ws.cell(r, cols["tier"]).value or "").strip().lower()
                     if "tier" in cols else None) or None
            if pid in decisions:
                errors.append(f"{tab} row {r}: {pid} appears in more than one tab")
                continue
            if raw == "":
                decisions[pid] = ("blank", tab, shown)
                stats["blank"] += 1
            elif raw not in DECISIONS:
                errors.append(f"{tab} row {r}: {pid}: DECISION {raw!r} is not in the "
                              f"vocabulary {sorted(DECISIONS)} and is not blank")
            elif raw == "unsure":
                decisions[pid] = ("unsure", tab, shown)
                stats["unsure"] += 1
            elif raw == "keep":
                decisions[pid] = ("keep", tab, shown)
                stats["keep"] += 1
            else:
                decisions[pid] = (raw, tab, shown)
                stats["change"] += 1
        per_tab[tab] = stats
    return decisions, per_tab, errors, no_tier_tabs


def fetch_store_tiers(env):
    """Current tier per provision from the store, so undecided rows ride
    through EXACTLY as stored (rule 1: unsure stays as it is in the store,
    and rows outside the review make no claim). Read-only. Returns None when
    no connection is available (dry-run fallback: base tiers, with a caveat)."""
    if not env.get("DATABASE_URL"):
        return None
    js = ("import pg from 'pg';"
          "const c = new pg.Client(process.env.DATABASE_URL); await c.connect();"
          "const r = await c.query('SELECT id, tier FROM standard_provision');"
          "console.log(JSON.stringify(r.rows)); await c.end();")
    res = subprocess.run(["node", "--input-type=module", "-e", js], env=env,
                         cwd=os.path.dirname(LOADER), capture_output=True, text=True)
    if res.returncode != 0:
        sys.exit(f"FAIL: could not read current store tiers (read-only query failed): {res.stderr.strip()[:200]}")
    return {row["id"]: row["tier"] for row in json.loads(res.stdout)}


def main():
    ap = argparse.ArgumentParser(description="Floor-review outputs -> loader, with the rules enforced")
    ap.add_argument("inputs", nargs="+",
                    help="One .xlsx workbook (session R), or one or more .tsv exports (session AN).")
    ap.add_argument("--base", default=DEFAULT_BASE)
    ap.add_argument("--out", default=os.path.join(HERE, "floor_review_merged_seed.json"),
                    help="Where the merged full seed JSON is written")
    ap.add_argument("--commit", action="store_true",
                    help="Run the loader (--supersede) after writing the merged seed. Default is dry run.")
    ap.add_argument("--allow-partial", action="store_true",
                    help="Proceed even though blank (unreviewed) rows remain. Never the default.")
    ap.add_argument("--flip-reviewed", action="store_true",
                    help="The separate explicit action: verify zero blank and zero unsure across all "
                         "three tabs, then run the loader with --reviewed. Refuses otherwise.")
    ap.add_argument("--accept-drift", action="store_true",
                    help="Session V: apply decided rows whose store tier differs from the tier the "
                         "workbook displayed. Never the default; the reviewer decided against "
                         "information that has since changed.")
    ap.add_argument("--reviewer",
                    help="Who performed the review. Required on any run that writes; recorded in "
                         "the frozen review record.")
    ap.add_argument("--record-out",
                    help="Where the review record is written (default: docs/ at the repo root).")
    args = ap.parse_args()

    if (args.commit or args.flip_reviewed) and not (args.reviewer or "").strip():
        sys.exit("FAIL: --reviewer is required on a run that writes; the review record names "
                 "who reviewed (the stamp decision, 2026-07-28).")

    with open(args.base) as f:
        base_rows = json.load(f)
    base_by_id = {r["provision_id"]: r for r in base_rows}

    exts = {os.path.splitext(p)[1].lower() for p in args.inputs}
    if exts <= {".tsv", ".txt"}:
        tsv_mode = True
        decisions, per_tab, errors, conflicts, notes = read_tsvs(args.inputs)
        no_tier_tabs = []
    elif exts == {".xlsx"} and len(args.inputs) == 1:
        tsv_mode = False
        conflicts, notes = [], {}
        decisions_raw, per_tab, errors, no_tier_tabs = read_workbook(args.inputs[0])
        # Unify: seen tiers are a LIST in AN mode; wrap the workbook's single.
        decisions = {pid: (d, tab, [shown] if shown else [])
                     for pid, (d, tab, shown) in decisions_raw.items()}
    else:
        sys.exit("FAIL: inputs must be a single .xlsx workbook or one or more .tsv exports, not a mix.")

    unknown = [pid for pid in decisions if pid not in base_by_id]
    errors.extend(f"{pid}: not in the base seed" for pid in unknown)
    if errors:
        print(f"FAIL: {len(errors)} input problems; nothing written.")
        for e in errors[:20]:
            print(f"    ! {e}")
        sys.exit(2)
    if conflicts:
        # Session AN: the inputs are not equivalent sources and the script
        # never arbitrates between them. The founder resolves each one.
        print(f"CONFLICT: {len(conflicts)} provision(s) carry different decisions in different "
              f"inputs. Nothing written; the founder resolves these, the script never picks:")
        for c in conflicts:
            print(f"    ! {c['provision_id']}: " + " vs ".join(c["values"]))
        sys.exit(2)
    if no_tier_tabs:
        print(f"NOTE: no tier column found on {', '.join(no_tier_tabs)}; the session V drift "
              f"check cannot run for decided rows on those tabs.")

    # The connection resolves early now: AN's completeness gate needs the
    # store's CURRENT tiers (which provisions are floors today), and the
    # drift check always did.
    env = dict(os.environ)
    if not env.get("DATABASE_URL"):
        neon = os.path.join(ROOT, ".neon-connection")
        if os.path.exists(neon):
            with open(neon) as f:
                env["DATABASE_URL"] = f.read().strip()
            print("DATABASE_URL read from .neon-connection (by name; not printed).")
    stored = fetch_store_tiers(env)
    if stored is None:
        if args.commit or args.flip_reviewed:
            sys.exit("FAIL: DATABASE_URL is not set and .neon-connection does not exist. "
                     "Set the variable or run from the machine that has the file.")
        print("CAVEAT (dry run, no connection): undecided rows shown with base-seed tiers; "
              "the commit run reads the store's current tiers instead.")
        stored = {}

    changes, keeps, unsure_rows = [], [], []
    shown_by_pid = {}
    blanks = 0
    for pid, (decision, tab, seen_list) in decisions.items():
        base = base_by_id[pid]
        shown_by_pid[pid] = seen_list
        if decision == "blank":
            blanks += 1
        elif decision == "unsure":
            unsure_rows.append({"provision_id": pid, "tier": base["tier"],
                                "text": base["text"][:100], "tab": tab})
        elif decision == "keep":
            keeps.append(pid)
        elif decision != base["tier"]:
            changes.append({"provision_id": pid, "from": base["tier"], "to": decision, "tab": tab})
        else:
            keeps.append(pid)  # decided tier equals current tier: confirmed, no move

    # Session AN completeness, against the provision set rather than tabs:
    # the flat export omits undecided rows entirely, so "no decision" is
    # computed from the store. A FLOOR provision without a decision blocks
    # the flip (the review's whole point was the floors); a non-floor row
    # the review never included is reported but does not block.
    if tsv_mode:
        current_tier = lambda pid: stored.get(pid) or base_by_id[pid]["tier"]
        missing = [pid for pid in base_by_id if pid not in decisions]
        floors_missing = [pid for pid in missing if current_tier(pid) in ("floor_1", "floor_2")]
        blanks = len(floors_missing)

    # The plan, always printed.
    print(f"Plan for {', '.join(os.path.basename(p) for p in args.inputs)} "
          f"against {len(base_rows)} base provisions:")
    for tab, s in per_tab.items():
        print(f"  {tab}: {s['rows']} rows; {s['change']} tier decisions, {s['keep']} keep, "
              f"{s['unsure']} unsure, {s['blank']} blank")
    print(f"  TOTAL: {len(changes)} tier changes, {len(keeps)} confirmed unchanged, "
          f"{len(unsure_rows)} unsure (never imported), {blanks} "
          + ("floor provisions with NO DECISION (the flip gate)" if tsv_mode
             else "blank (not yet reviewed)"))
    if tsv_mode:
        print(f"  No decision at all (not reviewed OR not included; the flat export cannot "
              f"tell them apart): {len(missing)} of {len(base_rows)}; "
              f"{len(floors_missing)} of those are floors, and floors gate the flip.")
        if floors_missing:
            print(f"    floors without a decision: {', '.join(sorted(floors_missing)[:15])}"
                  + (" ..." if len(floors_missing) > 15 else ""))
    for c in changes[:30]:
        print(f"    {c['provision_id']}: {c['from']} -> {c['to']}")
    if len(changes) > 30:
        print(f"    ... and {len(changes) - 30} more")
    if unsure_rows:
        print("  Unsure queue (stays exactly as stored):")
        for u in unsure_rows[:20]:
            print(f"    {u['provision_id']} [{u['tier']}] {u['text']}")

    # NOTE (session R report): there is no per-provision place to stamp who
    # reviewed a row and when. A keep produces no write, so at the end of the
    # review the only durable evidence is the store-level seed_reviewed
    # boolean. Adding one is its own reviewed change; see WORK_QUEUE.

    if args.flip_reviewed:
        if blanks or unsure_rows:
            what = ("floor provisions without a decision" if tsv_mode else "blank")
            print(f"REFUSED: seed_reviewed flips only at zero {what} and zero unsure; "
                  f"found {blanks} {what}, {len(unsure_rows)} unsure. The flag is a "
                  f"one-way door and turns once, when the review is complete.")
            sys.exit(2)
    elif blanks and not args.allow_partial:
        what = ("floor provisions carry no decision in any input"
                if tsv_mode else "rows are blank (not yet reviewed; blank is not keep)")
        print(f"REFUSED: {blanks} {what}. A partial "
              f"import must be explicit: re-run with --allow-partial, or finish the review.")
        sys.exit(2)

    # Merge over the FULL base seed so the loader's completeness check holds.
    # Only rows with an explicit tier decision take the review's value;
    # everything else (blank, unsure, outside the review) carries the STORE's
    # current tier, so an import can never silently revert a row the review
    # made no decision about.
    to_tier = {c["provision_id"]: c["to"] for c in changes}
    keep_set = set(keeps)

    # Session V: a decided row whose store tier differs from the tier the
    # review DISPLAYED was decided against stale information. Refuse those
    # rows unless the drift is explicitly accepted; never apply silently.
    # AN: a provision may carry several seen tiers (two agreeing inputs);
    # drift if ANY of them differs from the store.
    decided = [c["provision_id"] for c in changes] + keeps
    drifted = []
    for pid in decided:
        for shown in shown_by_pid.get(pid) or []:
            if pid in stored and stored[pid] != shown:
                drifted.append({"provision_id": pid, "shown": shown, "store": stored[pid]})
                break
    if drifted:
        print(f"DRIFT: {len(drifted)} decided row(s) were reviewed against a tier the store no "
              f"longer holds (changed after the workbook was built):")
        for d in drifted[:20]:
            print(f"    {d['provision_id']}: workbook showed {d['shown']}, store now {d['store']}")
        if not args.accept_drift:
            print("REFUSED: re-review those rows against the current tiers, or re-run with "
                  "--accept-drift to apply the decisions anyway. Silent application is the "
                  "failure this check exists to prevent (round seven, session V).")
            sys.exit(2)
        print("Drift accepted explicitly (--accept-drift); recorded in the review record.")
    # Fallback caveat when a tab carries no tier column: keeps compared to the
    # base sheet, the pre-V heuristic, still better than silence.
    drifted_keeps = [pid for pid in keeps
                     if not shown_by_pid.get(pid) and pid in stored
                     and stored[pid] != base_by_id[pid]["tier"]]
    if drifted_keeps:
        print(f"CAVEAT: {len(drifted_keeps)} keep row(s) (no tier column to compare) confirm the "
              f"sheet's tier but the store currently differs: {', '.join(drifted_keeps[:10])}")
    merged = []
    for r in base_rows:
        pid = r["provision_id"]
        if pid in to_tier:
            merged.append({**r, "tier": to_tier[pid]})
        elif pid in keep_set:
            # keep confirms the tier the reviewer SAW (the workbook's tier
            # column), falling back to the store's current tier when the tab
            # carries none. Never the base seed: the first proof run of this
            # very branch silently reverted a store row to its base tier,
            # which is the session V failure inside the fix. The drift check
            # above already refuses when seen and stored disagree.
            seen_list = shown_by_pid.get(pid) or []
            merged.append({**r, "tier": (seen_list[0] if seen_list else None) or stored.get(pid, r["tier"])})
        else:
            merged.append({**r, "tier": stored.get(pid, r["tier"])})
    with open(args.out, "w") as f:
        json.dump(merged, f, indent=1)
    print(f"Merged seed written: {args.out}")

    loader_args = ["node", "--experimental-strip-types", LOADER, args.out, "--supersede"]
    if args.flip_reviewed:
        loader_args.append("--reviewed")

    if not (args.commit or args.flip_reviewed):
        print("DRY RUN (default): the loader was not invoked. To apply:")
        print(f"  {' '.join(loader_args)}")
        return

    print(f"Invoking the loader ({'with --reviewed' if args.flip_reviewed else 'tier changes only'})...")
    res = subprocess.run(loader_args, env=env, cwd=os.path.dirname(LOADER))
    if res.returncode != 0:
        sys.exit(res.returncode)

    # The review stamp (founder decision 2026-07-28): a dated frozen record
    # is the durable evidence a keep otherwise never writes. The commit that
    # lands this file must add the printed hash line to frozen-records.test.ts,
    # which the frozen-records guard enforces in both directions.
    import hashlib
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)
    stamp = now.strftime("%Y-%m-%d %H:%M UTC")
    slug = now.strftime("%Y%m%d-%H%M")
    body_lines = [
        f"# Floor review record, {stamp}",
        "",
        f"Reviewer: {args.reviewer}. Input(s): "
        + ", ".join(os.path.basename(p) for p in args.inputs) + ".",
        f"Run mode: {'flip-reviewed (review complete, flag turned)' if args.flip_reviewed else 'tier import' + (' (partial, explicit)' if blanks else '')}.",
        "",
        "## Per input" if tsv_mode else "## Per tab",
        "",
    ]
    for tab, s in per_tab.items():
        body_lines.append(f"- {tab}: {s['rows']} rows; {s['change']} tier decisions, "
                          f"{s['keep']} keep, {s['unsure']} unsure, {s['blank']} blank")
    body_lines += ["", "## Tier changes applied", ""]
    # AN: the NOTE column carries the founder's policy and its date on the
    # never-do corrections; it stays in the durable record.
    body_lines += ([f"- {c['provision_id']}: {c['from']} to {c['to']}"
                    + (f" ({notes[c['provision_id']]})" if c["provision_id"] in notes else "")
                    for c in changes] or ["- none"])
    body_lines += ["", f"## Confirmed unchanged ({len(keeps)})", "",
                   ", ".join(sorted(keeps)) if keeps else "none"]
    body_lines += ["", f"## Unsure queue, never imported ({len(unsure_rows)})", ""]
    body_lines += ([f"- {u['provision_id']} [{u['tier']}]" for u in unsure_rows] or ["- none"])
    if tsv_mode:
        body_lines += ["", f"## No decision in any input: {len(missing)} of {len(base_rows)} "
                           f"({blanks} floors; floors gate the flip)"]
        if floors_missing:
            body_lines += ["", "Floors without a decision: " + ", ".join(sorted(floors_missing))]
    else:
        body_lines += ["", f"## Not yet reviewed (blank): {blanks}"]
    if drifted:
        body_lines += ["", "## Drift accepted explicitly", ""]
        body_lines += [f"- {d['provision_id']}: workbook showed {d['shown']}, store held {d['store']}"
                       for d in drifted]
    body = "\n".join(body_lines) + "\n"
    content = "---\nstatus: frozen\n---\n" + body
    record_dir = args.record_out or os.path.join(ROOT, "docs")
    record_path = os.path.join(record_dir, f"FLOOR_REVIEW_RECORD_{slug}.md")
    with open(record_path, "w") as f:
        f.write(content)
    digest = hashlib.sha256(body.encode()).hexdigest()
    rel = os.path.relpath(record_path, ROOT)
    if rel.startswith(".."):
        rel = f"docs/{os.path.basename(record_path)}"  # written elsewhere; manifest as if in docs/
    print(f"Review record written: {record_path}")
    print("Add this line to FROZEN in packages/schema/src/frozen-records.test.ts in the same "
          "commit that lands the record (the guard enforces it):")
    print(f'  "{rel}": "{digest}",')


if __name__ == "__main__":
    main()
