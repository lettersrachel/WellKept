#!/usr/bin/env python3
"""
wk_provisions.py : convert the founder's corrected provision review workbook
(WK_Provision_Seed_Founder_Review.xlsx) back into a provision seed JSON the
loader takes (packages/schema/src/load-provisions.ts), and report every row
the founder touched (CLAUDE_CODE_BRIEF T3).

The workbook carries provision_id, document, sec, ord, tier, pilot_default,
scope, text, and the founder columns I ("tier OK? (Y/change)") and J (notes).
Fields the sheet does not carry (kind, doc/section titles, effective_date,
version) join from the base seed by provision_id — the sheet is a review
overlay, not a re-extraction.

Column I semantics: "Y" = tier confirmed; a tier value = corrected tier
(applied); blank = unreviewed; anything else is flagged, not applied.
Tier assignments are POLICY (WK-DEV-005 S4): this script never invents one.

Usage:
  python3 wk_provisions.py REVIEW.xlsx --out corrected_seed.json
  python3 wk_provisions.py REVIEW.xlsx --out corrected_seed.json --report review_report.json
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl --break-system-packages")

TIERS = {"floor_1", "floor_2", "process", "method", "preference"}
DEFAULT_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "..", "seed", "provisions_seed.json")


def main():
    ap = argparse.ArgumentParser(description="Provision review workbook -> corrected seed JSON")
    ap.add_argument("workbook", help="The corrected founder review .xlsx")
    ap.add_argument("--base", default=DEFAULT_BASE,
                    help="Base provision seed JSON (fields the sheet does not carry)")
    ap.add_argument("--out", required=True, help="Write the corrected seed JSON here")
    ap.add_argument("--report", help="Write the founder-change report JSON here")
    args = ap.parse_args()

    with open(args.base) as f:
        base_by_id = {r["provision_id"]: r for r in json.load(f)}

    ws = openpyxl.load_workbook(args.workbook, data_only=True)["Provisions for review"]
    out_rows, errors = [], []
    changed, noted, unreviewed = [], [], []
    seen = set()
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        if not pid:
            continue
        pid = str(pid).strip()
        seen.add(pid)
        base = base_by_id.get(pid)
        if not base:
            errors.append(f"row {r}: {pid} not in base seed {args.base}")
            continue
        tier = str(ws.cell(r, 5).value or "").strip()
        founder = str(ws.cell(r, 9).value or "").strip()
        note = str(ws.cell(r, 10).value or "").strip()
        if founder and founder != "Y":
            if founder in TIERS:
                changed.append({"provision_id": pid, "from": tier, "to": founder,
                                "note": note})
                tier = founder
            else:
                errors.append(f"row {r}: {pid}: column I is {founder!r} — not Y and "
                              f"not a tier; resolve before load")
        elif not founder:
            unreviewed.append(pid)
        if note and founder in ("Y", ""):
            noted.append({"provision_id": pid, "note": note})
        if tier not in TIERS:
            errors.append(f"row {r}: {pid}: tier {tier!r} not in {sorted(TIERS)}")
            continue
        text = str(ws.cell(r, 8).value or "").strip()
        if not text:
            errors.append(f"row {r}: {pid}: empty text")
            continue
        scope = [s.strip() for s in str(ws.cell(r, 7).value or "").split(";") if s.strip()]
        out_rows.append({**base,
                         "tier": tier,
                         "pilot_default": str(ws.cell(r, 6).value or "").strip() == "YES",
                         "scope": scope or base["scope"],
                         "text": text})

    missing = sorted(set(base_by_id) - seen)
    dupes = len(out_rows) - len({r["provision_id"] for r in out_rows})
    if dupes:
        errors.append(f"{dupes} duplicate provision_id rows in the sheet")
    if missing:
        errors.append(f"{len(missing)} base provisions missing from the sheet: "
                      f"{', '.join(missing[:10])}")
    if errors:
        print(f"FAIL: {len(errors)} problems; nothing written.")
        for e in errors[:20]:
            print(f"    ! {e}")
        sys.exit(2)

    with open(args.out, "w") as f:
        json.dump(out_rows, f, indent=1)
    report = {
        "workbook": args.workbook,
        "rows": len(out_rows),
        "tier_changes": changed,          # column I != Y: the founder's corrections
        "notes": noted,                   # column J on otherwise-confirmed rows
        "unreviewed": len(unreviewed),    # column I blank
        "unreviewed_ids": unreviewed[:50],
    }
    if args.report:
        with open(args.report, "w") as f:
            json.dump(report, f, indent=1)
    print(f"Wrote {len(out_rows)} provisions to {args.out}")
    print(f"Founder review: {len(changed)} tier changes, {len(noted)} notes, "
          f"{len(unreviewed)} unreviewed rows"
          + (f" (report: {args.report})" if args.report else ""))
    if unreviewed:
        print("NOTE: unreviewed rows load with their extracted tiers; the "
              "seed_reviewed flag stays false until the review is complete.")


if __name__ == "__main__":
    main()
