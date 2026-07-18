#!/usr/bin/env python3
"""
wk_import.py : Well Kept intake workbook importer (US-01, WK-APP-003 S5)

Parses a WK_PLAY_002 Intake Workbook (.xlsx) into the seed schema the
application imports: one field record per workbook row, carrying section,
name, value, provenance, provenanceDate, sensitivity, flag, and note.

Modes:
  Dry run (default): parse and report, and if --against is given, show the
  full diff (added / removed / changed) without writing anything.
  Commit (--commit PATH): write the seed JSON.

Usage:
  python3 wk_import.py WORKBOOK.xlsx --household "Fernbrook" --tier family_ops
  python3 wk_import.py WORKBOOK.xlsx --against wk_seed_intake_schema.json
  python3 wk_import.py WORKBOOK.xlsx --against seed.json --commit new_seed.json

Rules enforced (WK-DEV-005):
  Blank means unasked; "N/A-confirmed" is a VALUE and is preserved verbatim,
  never coalesced. Sensitivity fails closed: an unrecognized marker maps UP.
  UUIDs are preserved from --against where section+name match exactly, so a
  re-import never churns identities.

No real S3 values belong in any file this script touches during the pilot's
demo phase; the vault is the Year 2 build's job.
"""

import argparse
import json
import re
import sys
import uuid

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl --break-system-packages")

SENS_ORDER = {"s1": 1, "s2": 2, "s3": 3}


def infer_sensitivity(name: str, sens_cell) -> str:
    """Explicit Sens. column wins; else markers in the field text; else s1.
    Anything unrecognized or S3-adjacent maps UP (fail closed)."""
    if sens_cell:
        raw = str(sens_cell).strip().lower()
        if raw in SENS_ORDER:
            return raw
        if "3" in raw:
            return "s3"
        if "2" in raw:
            return "s2"
        if "1" in raw:
            return "s1"
        return "s3"  # unrecognized explicit marking: fail closed
    m = re.search(r"\[(S[123][^\]]*)\]", name, re.IGNORECASE)
    if m:
        tag = m.group(1).lower()
        if tag.startswith("s3"):
            return "s3"
        if tag.startswith("s2"):
            return "s2"
        return "s1"
    low = name.lower()
    if any(k in low for k in ("alarm code", "lockbox", "safe combination", "firearm",
                              "gate code", "lock combination", "key location", "keypad",
                              "door code", "combination")):
        return "s3"
    if any(k in low for k in ("clearance", "classified", "medication", "controlled substance",
                              "custody", "religious", "immigration", "diagnosis")):
        return "s2"
    return "s1"


def parse_workbook(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    fields = []
    tabs_read = []
    blank_sens = []
    for ws in wb.worksheets:
        title = ws.title
        if not title.startswith("S") or "Intelligence" in title:
            continue
        tabs_read.append(title)
        for r in range(1, ws.max_row + 1):
            sec = ws.cell(r, 1).value
            name = ws.cell(r, 2).value
            if not isinstance(sec, (int, float)) or not name or not str(name).strip():
                continue
            name = str(name).strip()
            value = ws.cell(r, 3).value
            if not ws.cell(r, 6).value:
                blank_sens.append((int(sec), name))
            confirmed_by = ws.cell(r, 4).value
            date = ws.cell(r, 5).value
            sens_cell = ws.cell(r, 6).value
            flag = ws.cell(r, 7).value
            note = ws.cell(r, 8).value

            value_s = "" if value is None else str(value).strip()
            # Canonical provenance tokens per WK-DEV-005 glossary (underscores).
            raw_prov = str(confirmed_by).strip() if confirmed_by else "unconfirmed"
            prov = raw_prov.lower().replace("-", "_").replace(" ", "_")
            if prov not in ("asked", "observed", "verified_by_touch", "client_written",
                            "n/a_confirmed", "na_confirmed", "unconfirmed"):
                prov = raw_prov  # unknown token preserved verbatim, never coalesced
            fields.append({
                "section": int(sec),
                "name": name,
                "value": value_s,
                "provenance": prov,
                "provenanceDate": str(date).strip() if date else "",
                "sensitivity": infer_sensitivity(name, sens_cell),
                "confirmed": bool(value_s) and prov != "unconfirmed",
                "flag": str(flag).strip() if flag else "none",
                "note": str(note).strip() if note else "",
            })
    return fields, tabs_read, blank_sens


def key(f):
    return (f["section"], f["name"])


def diff(new_fields, old):
    old_by_key = {key(f): f for f in old.get("fields", [])}
    new_by_key = {key(f): f for f in new_fields}
    added = [k for k in new_by_key if k not in old_by_key]
    removed = [k for k in old_by_key if k not in new_by_key]
    changed = []
    for k in new_by_key:
        if k in old_by_key:
            a, b = old_by_key[k], new_by_key[k]
            deltas = [c for c in ("value", "provenance", "provenanceDate", "sensitivity", "flag", "note")
                      if str(a.get(c, "")) != str(b.get(c, ""))]
            if deltas:
                changed.append((k, deltas, a, b))
    return added, removed, changed, old_by_key


def main():
    ap = argparse.ArgumentParser(description="Well Kept intake workbook importer")
    ap.add_argument("workbook", help="Path to the WK_PLAY_002 intake workbook (.xlsx)")
    ap.add_argument("--household", default="Unnamed Household")
    ap.add_argument("--tier", default="family_ops",
                    choices=["essential", "family_ops", "concierge"])
    ap.add_argument("--against", help="Existing seed JSON to diff against (preserves UUIDs)")
    ap.add_argument("--commit", help="Write seed JSON to this path (otherwise dry run)")
    ap.add_argument("--template", action="store_true",
                    help="Template mode: permit sensitivity inference for blank Sens. cells. "
                         "Without this flag (real household imports), rows with a blank Sens. "
                         "column FAIL LOUDLY per WK-DEV-005 Section 3.")
    args = ap.parse_args()

    fields, tabs, blank_sens = parse_workbook(args.workbook)
    print(f"Parsed {len(fields)} field rows from {len(tabs)} tabs: {', '.join(tabs)}")
    if blank_sens and not args.template:
        print(f"\nFAIL (WK-DEV-005 S3): {len(blank_sens)} rows have a blank Sens. column. "
              f"Real imports must not infer sensitivity. Fill the column, or re-run with "
              f"--template if this is the blank template.")
        for sec, name in blank_sens[:15]:
            print(f"    ! S{sec}  {name[:75]}")
        sys.exit(2)

    old = None
    if args.against:
        with open(args.against) as f:
            old = json.load(f)
        added, removed, changed, old_by_key = diff(fields, old)
        print(f"\nDRY-RUN DIFF against {args.against}")
        print(f"  added:   {len(added)}")
        for k in sorted(added)[:20]:
            print(f"    + S{k[0]}  {k[1][:80]}")
        print(f"  removed: {len(removed)}")
        for k in sorted(removed)[:20]:
            print(f"    - S{k[0]}  {k[1][:80]}")
        print(f"  changed: {len(changed)}")
        for k, deltas, a, b in changed[:20]:
            print(f"    ~ S{k[0]}  {k[1][:60]}  [{', '.join(deltas)}]")
        if not (added or removed or changed):
            print("  Clean: workbook and seed agree exactly.")

    # assemble output (UUIDs preserved where possible)
    preserved = created = 0
    out_fields = []
    old_by_key = {key(f): f for f in (old.get("fields", []) if old else [])}
    for f in fields:
        prior = old_by_key.get(key(f))
        if prior:
            fid = prior["id"]
            preserved += 1
            # Fail closed: sensitivity never downgrades below the prior record.
            if SENS_ORDER.get(prior.get("sensitivity", "s1"), 1) > SENS_ORDER.get(f["sensitivity"], 1):
                f["sensitivity"] = prior["sensitivity"]
        else:
            fid = str(uuid.uuid4())
            created += 1
        out_fields.append({"id": fid, **f})

    out = {
        "household": {"id": (old or {}).get("household", {}).get("id") or str(uuid.uuid4()),
                      "name": args.household, "tier": args.tier},
        "_meta": {
            "importedFrom": args.workbook,
            "importer": "wk_import.py v1.0",
            "fieldCount": len(out_fields),
            "uuids": {"preserved": preserved, "created": created},
        },
        "fields": out_fields,
    }

    if args.commit:
        with open(args.commit, "w") as f:
            json.dump(out, f, indent=1)
        print(f"\nCOMMITTED: {args.commit} ({len(out_fields)} fields, "
              f"{preserved} UUIDs preserved, {created} new)")
    else:
        print(f"\nDry run only. {preserved} UUIDs would be preserved, {created} created. "
              f"Re-run with --commit PATH to write.")


if __name__ == "__main__":
    main()
